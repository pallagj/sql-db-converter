import streamlit as st
from openpyxl import Workbook
from sqlalchemy import inspect
import yaml
import streamlit as st
import re
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

st.set_page_config(layout="centered", page_title="To excel", page_icon="➡️")
st.markdown("""
    <style>
    div[data-testid='stSidebarNav'] ul {max-height:none}</style>
    """, unsafe_allow_html=True)

try:
    with open('config.yaml', 'r') as file:
        config = yaml.safe_load(file)
except Exception as e:
    config = {'transitions': {}}


def has_keys():
    return all(key in config for key in ['transitions', 'filter', 'connection'])


if 'connection' in st.session_state and has_keys():
    st.sidebar.caption(f"🔗 {st.session_state.db_name.capitalize()}")
    engine = st.session_state.connection.engine

    st.subheader("Convert to excel")
    file_name = st.text_input("Excel file name", "output.xlsx")

    table_filter = config.get('filter').get('table_filter')
    tables = config.get('filter').get('tables')

    field_filter = config.get('filter').get('field_filter')
    fields = config.get('filter').get('fields')

    transitions = config.get("transitions")

    if st.button("Convert"):
        try:
            inspector = inspect(engine)
            wb = Workbook()
            ws = wb.active
            wb.remove(ws)
            for table in inspector.get_table_names():
                if (table_filter == 'include') == any(re.fullmatch(t, table) for t in tables):
                    # Get the columns for the table
                    columns = inspector.get_columns(table)
                    query_fields = []
                    for column in columns:
                        if (field_filter == 'include') == any(re.fullmatch(f, column['name']) for f in fields):
                            # Get type of the column
                            col_type = str(column['type']).upper()
                            # Split the type and length if exists
                            col_type, parentheses, col_length = re.match(r'(\w+)(\((.+)\))?', col_type).groups('')
                            # Translate the type using the yaml file
                            col_type_translated = transitions.get(col_type, col_type)
                            # Append the translated type to the column name
                            query_fields.append(f"`{column['name']}` as `{column['name'].replace('_', ' ').capitalize()} ({col_type_translated + (' - max ' if col_length else '') + col_length})`")
                    # Build and run the query
                    query = f"SELECT {', '.join(query_fields)} FROM `{table}`"
                    df = pd.read_sql_query(query, engine)
                    df = df.astype(str)  # convert all columns to string type
                    ws = wb.create_sheet(title=table)
                    # for row in dataframe_to_rows(df, index=False, header=True):
                    #    ws.append(row)
                    # ws = wb.create_sheet(title=table)
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                        ws.append(row)
                        if r_idx == 1:  # header row
                            for c_idx, value in enumerate(row, start=1):  # adjust columns width
                                column = get_column_letter(c_idx)
                                max_length = 0
                                column = ws.column_dimensions[column]
                                if len(str(value)) > max_length:
                                    max_length = len(str(value))
                                column.width = max_length
            wb.save(file_name)
            st.success(f'The data has been successfully exported to {file_name}')
        except Exception as e:
            st.error(f'Failed to run query: {e}')
else:
    st.error("There is no database connection!")
